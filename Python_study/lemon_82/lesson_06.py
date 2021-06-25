# python利用Excel读取和存储数据
import requests
import openpyxl
from openpyxl import load_workbook

#定义读取Excel的函数
def read_excel(excel_name,sheet_name) :
    cases = []
    wb = load_workbook(excel_name)
    sheet = wb[sheet_name]
    max_r = sheet.max_row  #获取最大行号
    for i in range(2,max_r+1) :
        case = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i, column=6).value,
        except_result = sheet.cell(row=i, column=7).value)
        cases.append(case)
    return cases

#定义接口函数：
def api_requests(url,data) :
    head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    response = requests.post(url=url,json=data,headers=head)
    return response.json()

#定义写入Excel的函数
def write_excel(excel_name,sheet_name,row_number,final_result) :
    wb = load_workbook(excel_name)
    sheet = wb[sheet_name]
    cell = sheet.cell(row=row_number,column=8)
    cell.value = final_result
    wb.save("testcase_api_wuye.xlsx")

def execute(excel_name,sheet_name) :
    cases = read_excel(excel_name,sheet_name)
    for case in cases :
        case_id = int(case["case_id"])
        url = case["url"]
        data = eval(case["data"])
        except_result = eval(case["except_result"])
        actually_result = api_requests(url,data)
        if actually_result["msg"] == except_result["msg"] :
            final_result = "PASS"
        else :
            final_result = "FAIL"
        write_excel(excel_name,sheet_name,case_id+1,final_result)

execute("testcase_api_wuye.xlsx","register")

