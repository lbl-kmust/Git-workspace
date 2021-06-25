import requests
import pprint
import jsonpath
# 注册接口：
# url_register = "http://47.115.15.198:7001/smarthome/user/register"
# data_register = {
#     "phone":"15950542236",
#     "pwd":"1234567a","rePwd":"1234567a",
#     "userName":"柠檬可爱",
#     "verificationCode":"611203"}
# head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
# response = requests.post(url=url_register,json=data_register,headers=head)
# pprint.pprint(response.json())

# 登录接口：
# url_login = "http://47.115.15.198:7001/smarthome/user/login"
# data_login = {"pwd": "1234567a","userName": "柠檬可爱"}
# head = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
# response_login = requests.post(url=url_login,json=data_login,headers=head)  #关键字传参
# pprint.pprint(response_login.json())
# # 字典取值：
# # user_id = response_login.json()["data"]["id"]  #关键字取值
# # user_token = response_login.json()["data"]["token_info"]["token"]
# user_id = jsonpath.jsonpath(response_login.json(),"$..id")[0] #jsonpath取出的值是保存在列表里面
# user_token = jsonpath.jsonpath(response_login.json(),"$..token")[0]
# # 完善信息接口：
# url_complete = "http://47.115.15.198:7001/smarthome/merchant/complete"
# data_complete = {
# "address": "云南省昆明市罗白馨苑",
# "establishDate": "2021-04-02",
# "legalPerson": "孙悟空",
# "licenseCode": "123456789012345",
# "licenseUrl": "http://127.0.0.1/smarthome/aaa.jpg",
# "merchantName": "蓝莓科技有限公司",
# "merchantType": 2,
# "registerAuthority": "安宁区派出所",
# "tel": "13888888888",
# "userId": user_id,
# "validityDate": "2033-05-02"}
# head_complete = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json","Authorization":"Bearer "+user_token}
# response_complete = requests.put(url=url_complete,json=data_complete,headers=head_complete)  #关键字传参
# print(response_complete.json())

# 获取一个百度页面的请求
# url_baidu = "https://www.baidu.com/"
# head_baidu = {"User-Agent":
# 	"Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) Gecko/20100101 Firefox/88.0"}
# response_baidu = requests.get(url=url_baidu,headers=head_baidu)
# print(response_baidu.content.decode("utf8"))

# 获取一个带参数百度页面的请求
# url_baidu = "https://www.baidu.com/s"
# para_baidu = {"wd":"欧洲杯"}
# head_baidu = {"User-Agent":
# 	"Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) Gecko/20100101 Firefox/88.0"}
# response_baidu = requests.get(url=url_baidu,params=para_baidu,headers=head_baidu)
# print(response_baidu.content.decode("utf8"))

import openpyxl
from openpyxl import load_workbook
wb = load_workbook("testcase_api_wuye.xlsx") #加载工作簿
sheet = wb["login"] #选择表单
cell = sheet.cell(row=2,column=5) #选择行列值
print(cell.value)